#!/usr/bin/env python3
"""
gerar_relatorio.py

Gera:
  html_final.html      - Calendario (week-lines) + Heatmap de vendas por categoria
  relatorio_vendas.xlsx - Dados consolidados de todos os canais

Fonte: E:\\arquivos google\\resultado_v2\\<canal>\\
  resultado_unicos_*.csv  -> calendario (nome_header == PERIODO)
  resultado_filtrado_*.csv -> heatmap + excel

Uso: python gerar_relatorio.py
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import os
import re
from collections import defaultdict
from typing import Any, Dict, List, Optional, Tuple

HERE         = os.path.dirname(os.path.abspath(__file__))
BASE_DIR     = os.path.join(HERE, "input")
OUTPUT_HTML  = os.path.join(HERE, "output", "html_final.html")
OUTPUT_EXCEL = os.path.join(HERE, "output", "relatorio_vendas.xlsx")

PALETTE = [
    '#1976d2', '#e53935', '#43a047', '#fb8c00', '#8e24aa',
    '#00acc1', '#f4511e', '#6d4c41', '#3949ab', '#d81b60',
    '#00897b', '#f9a825', '#546e7a', '#26a69a', '#c62828',
]

DATE_PAT     = re.compile(r'^(\d{4}-\d{2}-\d{2})')
FILE_DATE_RE = re.compile(r'(\d{8})(?:\.csv)?$', re.IGNORECASE)


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_float(s: Any) -> float:
    try:
        return float(str(s).strip().strip('"').replace(',', '.'))
    except (ValueError, AttributeError):
        return 0.0


def file_date_key(filename: str) -> int:
    m = FILE_DATE_RE.search(os.path.basename(filename))
    return int(m.group(1)) if m else 0


def fmt_channel(name: str) -> str:
    return name.replace('_', ' ').title()


def fmt_file_date(filename: str) -> str:
    m = FILE_DATE_RE.search(filename)
    if m:
        try:
            return dt.datetime.strptime(m.group(1), '%Y%m%d').strftime('%d/%m/%Y')
        except ValueError:
            pass
    return os.path.splitext(filename)[0]



# ── Leitura resultado_unicos (calendario) ─────────────────────────────────────

def read_unicos(canal_dir: str) -> Dict[str, List[str]]:
    """Retorna { nome_arquivo: [segundas-feiras ISO] }"""
    result: Dict[str, set] = defaultdict(set)
    for fname in sorted(os.listdir(canal_dir)):
        if 'resultado_unicos' not in fname.lower() or not fname.lower().endswith('.csv'):
            continue
        fpath = os.path.join(canal_dir, fname)
        try:
            with open(fpath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get('nome_header', '').strip() != 'PERIODO':
                        continue
                    val = row.get('valor', '').strip()
                    m = DATE_PAT.match(val)
                    if not m:
                        continue
                    nome_arq = row.get('nome_arquivo', '').strip()
                    if nome_arq:
                        result[nome_arq].add(m.group(1))
        except Exception as e:
            print(f'    [Aviso unicos] {fname}: {e}')
    return {k: sorted(v) for k, v in result.items()}


# ── Leitura resultado_filtrado (heatmap + excel) ──────────────────────────────

def read_filtrado_canal(canal: str, canal_dir: str) -> List[Dict]:
    """
    Le todos resultado_filtrado_*.csv do canal.
    Deduplicacao por (nome_arquivo, PERIODO, DESCRICAO_PRODUTO):
    prevalece o arquivo resultado_filtrado mais recente (data no nome).
    Retorna lista de dicts.
    """
    filtrado_files: List[Tuple[int, str]] = []
    for fname in os.listdir(canal_dir):
        if 'resultado_filtrado' not in fname.lower() or not fname.lower().endswith('.csv'):
            continue
        filtrado_files.append((file_date_key(fname), fname))
    filtrado_files.sort()  # mais antigo primeiro; ultimo vence na dedup

    dedup: Dict[Tuple, Dict] = {}
    for _, fname in filtrado_files:
        fpath = os.path.join(canal_dir, fname)
        try:
            with open(fpath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    periodo_raw = row.get('PERIODO', '').strip()
                    m = DATE_PAT.match(periodo_raw)
                    if not m:
                        continue
                    periodo  = m.group(1)
                    nome_arq = row.get('nome_arquivo', '').strip()
                    descricao = row.get('DESCRICAO_PRODUTO', '').strip()
                    if not (nome_arq and descricao):
                        continue
                    key = (nome_arq, periodo, descricao)
                    dedup[key] = {
                        'canal':             canal,
                        'arquivo_filtrado':  fname,
                        'nome_arquivo':      nome_arq,
                        'PERIODO':           periodo,
                        'DESCRICAO_PRODUTO': descricao,
                        'DN_TOTAL':          parse_float(row.get('DN_TOTAL', 0)),
                        'DP_TOTAL':          parse_float(row.get('DP_TOTAL', 0)),
                        'VENDAS_VALOR':      parse_float(row.get('VENDAS_VALOR', 0)),
                        'VENDAS_VOLUME':     parse_float(row.get('VENDAS_VOLUME_gr_ml', 0)),
                        'VENDAS_UNITARIAS':  parse_float(row.get('VENDAS_UNITARIAS', 0)),
                    }
        except Exception as e:
            print(f'    [Aviso filtrado] {fname}: {e}')
    return list(dedup.values())


# ── Build SALES_DATA ──────────────────────────────────────────────────────────

def build_sales_data(all_rows: List[Dict]) -> Dict:
    """
    Retorna { canal: { descricao: { periodo: {v, vol, u} } } }
    Valores sao somados por (canal, descricao, periodo) para agregar
    diferentes arquivos do mesmo canal.
    """
    data: Any = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {'v': 0.0, 'vol': 0.0, 'u': 0.0}
    )))
    for row in all_rows:
        cell = data[row['canal']][row['DESCRICAO_PRODUTO']][row['PERIODO']]
        cell['v']   += row['VENDAS_VALOR']
        cell['vol'] += row['VENDAS_VOLUME']
        cell['u']   += row['VENDAS_UNITARIAS']

    # Arredondamento e conversao para dict simples
    result = {}
    for canal, cats in data.items():
        result[canal] = {}
        for desc, weeks in cats.items():
            result[canal][desc] = {}
            for week, vals in weeks.items():
                result[canal][desc][week] = {
                    'v':   round(vals['v'],   2),
                    'vol': round(vals['vol'], 2),
                    'u':   round(vals['u'],   0),
                }
    return result


# ── Excel ─────────────────────────────────────────────────────────────────────

def write_excel(all_rows: List[Dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  [Aviso] openpyxl nao instalado. Excel nao gerado.')
        return

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas'

    cols = [
        ('canal',             'canal'),
        ('arquivo_filtrado',  'arquivo_filtrado'),
        ('nome_arquivo',      'nome_arquivo'),
        ('PERIODO',           'PERIODO'),
        ('DESCRICAO_PRODUTO', 'DESCRICAO_PRODUTO'),
        ('DN_TOTAL',          'DN_TOTAL'),
        ('DP_TOTAL',          'DP_TOTAL'),
        ('VENDAS_VALOR',      'VENDAS_VALOR'),
        ('VENDAS_VOLUME_gr_ml', 'VENDAS_VOLUME'),
        ('VENDAS_UNITARIAS',  'VENDAS_UNITARIAS'),
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')

    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    rows_sorted = sorted(all_rows, key=lambda r: (
        r['canal'], r['nome_arquivo'], r['PERIODO'], r['DESCRICAO_PRODUTO']
    ))

    for row_idx, row in enumerate(rows_sorted, 2):
        for col_idx, (_, key) in enumerate(cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    # Ajusta larguras
    col_widths = [len(h) for h, _ in cols]
    for row in rows_sorted[:2000]:
        for i, (_, key) in enumerate(cols):
            v = str(row.get(key, ''))
            col_widths[i] = max(col_widths[i], min(len(v), 45))
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w + 2

    wb.save(OUTPUT_EXCEL)
    print(f'Excel gerado em: {OUTPUT_EXCEL}')


# ── Template HTML ─────────────────────────────────────────────────────────────
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Relatorio FTP 2025-2026</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    :root {
      --blue:       #1976d2;
      --blue-dark:  #1255a0;
      --blue-pale:  #e3f2fd;
      --blue-hover: #bbdefb;
      --sidebar-w:  260px;
      --text:       #1e2a3a;
      --text-sub:   #607080;
      --weekend:    #8faabf;
      --border:     #e0eaf4;
      --card-bg:    #ffffff;
      --body-bg:    #f0f5fb;
      --radius:     14px;
      --shadow:     0 1px 3px rgba(21,101,192,.08), 0 4px 16px rgba(21,101,192,.07);
      --shadow-h:   0 6px 24px rgba(21,101,192,.16);
    }

    html, body { height: 100%; }
    body {
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
      background: var(--body-bg);
      color: var(--text);
      display: flex;
      min-height: 100vh;
      overflow: hidden;
    }

    /* ── SIDEBAR ── */
    .sidebar {
      width: var(--sidebar-w); min-width: var(--sidebar-w);
      height: 100vh; background: var(--card-bg);
      border-right: 1px solid var(--border);
      box-shadow: 2px 0 16px rgba(21,101,192,.10);
      display: flex; flex-direction: column;
      position: sticky; top: 0; z-index: 20;
    }
    .sidebar-brand { padding: 22px 20px 18px; border-bottom: 1px solid var(--border); flex-shrink: 0; }
    .brand-text h2 { font-size: .95rem; font-weight: 700; color: var(--blue-dark); line-height: 1.2; }
    .brand-text span { font-size: .7rem; color: var(--text-sub); }

    .sidebar-body {
      flex: 1; overflow-y: auto; padding: 12px 0 8px;
      scrollbar-width: thin; scrollbar-color: var(--border) transparent;
    }
    .sidebar-body::-webkit-scrollbar { width: 4px; }
    .sidebar-body::-webkit-scrollbar-thumb { background: var(--border); border-radius: 4px; }

    .sb-section {
      font-size: .65rem; font-weight: 700; letter-spacing: .10em;
      text-transform: uppercase; color: var(--text-sub); padding: 10px 20px 6px; opacity: .8;
    }
    .sb-channel {
      display: flex; align-items: center; gap: 10px; padding: 10px 20px;
      font-size: .875rem; font-weight: 500; color: var(--text-sub); cursor: pointer;
      transition: background .13s, color .13s, border-color .13s;
      border-left: 3px solid transparent; user-select: none;
    }
    .sb-channel:hover { background: var(--blue-pale); color: var(--blue); border-left-color: var(--blue-hover); }
    .sb-channel.active { background: var(--blue-pale); color: var(--blue); font-weight: 600; border-left-color: var(--blue); }
    .ch-dot { width: 8px; height: 8px; border-radius: 50%; background: #c8d8e8; flex-shrink: 0; transition: background .13s; }
    .sb-channel:hover .ch-dot, .sb-channel.active .ch-dot { background: var(--blue); }
    .ch-count { margin-left: auto; font-size: .7rem; font-weight: 600; background: var(--blue-hover); color: var(--blue); border-radius: 10px; padding: 1px 7px; }

    .sb-back {
      display: flex; align-items: center; gap: 8px; padding: 10px 16px 12px;
      font-size: .8rem; font-weight: 600; color: var(--blue); cursor: pointer;
      border-bottom: 1px solid var(--border); margin-bottom: 4px; transition: background .13s;
    }
    .sb-back:hover { background: var(--blue-pale); }
    .sb-back svg { width: 16px; height: 16px; fill: currentColor; }
    .sb-chan-header { padding: 8px 20px 10px; font-size: .85rem; font-weight: 700; color: var(--blue-dark); }
    .sb-file {
      display: flex; align-items: center; gap: 10px; padding: 7px 16px;
      font-size: .8rem; font-weight: 500; color: var(--text-sub); cursor: pointer;
      border-radius: 6px; margin: 1px 8px; transition: background .12s; user-select: none;
    }
    .sb-file:hover { background: var(--blue-pale); }
    .sb-file.off { opacity: .4; }
    .file-swatch { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }
    .file-label { flex: 1; }
    .file-check {
      width: 14px; height: 14px; border-radius: 3px; border: 2px solid #c8d8e8;
      background: #fff; display: flex; align-items: center; justify-content: center;
      flex-shrink: 0; transition: border-color .12s, background .12s;
    }
    .file-check.on { border-color: var(--blue); background: var(--blue); }
    .file-check.on::after {
      content: ''; display: block; width: 4px; height: 7px;
      border: 2px solid #fff; border-top: none; border-left: none;
      transform: rotate(45deg) translate(-1px, -1px);
    }
    .sb-controls { display: flex; gap: 6px; padding: 6px 16px 4px; }
    .sb-btn {
      font-size: .7rem; font-weight: 600; padding: 3px 10px; border-radius: 20px;
      cursor: pointer; border: 1px solid var(--border); background: var(--card-bg);
      color: var(--text-sub); transition: background .12s, color .12s; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    }
    .sb-btn:hover { background: var(--blue-pale); color: var(--blue); border-color: var(--blue-hover); }
    .sb-div { height: 1px; background: var(--border); margin: 8px 14px; }
    .sidebar-footer {
      padding: 12px 20px; border-top: 1px solid var(--border);
      font-size: .72rem; color: var(--text-sub); background: #fafcff; flex-shrink: 0;
    }
    .sidebar-footer strong { display: block; font-weight: 600; color: var(--text); margin-bottom: 1px; }

    /* ── MAIN ── */
    .main {
      flex: 1; height: 100vh; overflow-y: auto; padding: 36px 32px 56px;
      scrollbar-width: thin; scrollbar-color: var(--border) transparent;
    }
    .main::-webkit-scrollbar { width: 6px; }
    .main::-webkit-scrollbar-thumb { background: var(--border); border-radius: 6px; }

    .page-header { margin-bottom: 24px; }
    .page-header h1 { font-size: 1.55rem; font-weight: 700; color: var(--blue-dark); letter-spacing: -.015em; }
    .page-header p { margin-top: 5px; font-size: .88rem; color: var(--text-sub); }
    .today-chip {
      display: inline-flex; align-items: center; gap: 6px; margin-top: 10px;
      background: var(--blue-pale); border: 1px solid var(--blue-hover);
      border-radius: 20px; padding: 4px 12px 4px 8px;
      font-size: .78rem; font-weight: 600; color: var(--blue);
    }
    .today-chip .chip-dot { width: 8px; height: 8px; background: var(--blue); border-radius: 50%; }

    /* ── TABS ── */
    .tab-bar { display: flex; gap: 4px; margin-bottom: 24px; border-bottom: 2px solid var(--border); }
    .tab-btn {
      padding: 10px 20px; font-size: .875rem; font-weight: 600;
      color: var(--text-sub); background: none; border: none;
      border-bottom: 2px solid transparent; margin-bottom: -2px;
      cursor: pointer; transition: color .13s, border-color .13s; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    }
    .tab-btn:hover { color: var(--blue); }
    .tab-btn.active { color: var(--blue); border-bottom-color: var(--blue); }

    /* ── CALENDARIO ── */
    .year-section { margin-bottom: 44px; }
    .year-heading { display: flex; align-items: center; gap: 14px; margin-bottom: 22px; }
    .year-num { font-size: 1.05rem; font-weight: 700; color: var(--blue); letter-spacing: .04em; flex-shrink: 0; }
    .year-line { flex: 1; height: 2px; background: linear-gradient(90deg, var(--blue-hover) 0%, transparent 100%); border-radius: 2px; }
    .months-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(252px, 1fr)); gap: 18px; }
    .cal-card {
      background: var(--card-bg); border-radius: var(--radius);
      box-shadow: var(--shadow); border: 1px solid var(--border);
      padding: 18px 14px 14px; transition: box-shadow .18s, transform .18s;
    }
    .cal-card:hover { box-shadow: var(--shadow-h); transform: translateY(-1px); }
    .cal-card.is-current { border-color: var(--blue-hover); box-shadow: 0 0 0 2px var(--blue-pale), var(--shadow-h); }
    .cal-card-top { display: flex; align-items: center; justify-content: space-between; margin-bottom: 14px; }
    .cal-month-label { font-size: .9rem; font-weight: 700; color: var(--blue-dark); }
    .cal-badge { font-size: .67rem; font-weight: 700; letter-spacing: .06em; text-transform: uppercase; color: var(--blue); background: var(--blue-pale); border-radius: 20px; padding: 3px 9px; }
    .cal-dow-row { display: grid; grid-template-columns: repeat(7, 1fr); margin-bottom: 4px; }
    .cal-dow-cell { text-align: center; font-size: .62rem; font-weight: 700; letter-spacing: .05em; text-transform: uppercase; padding: 3px 0 5px; }
    .cal-dow-cell:nth-child(1), .cal-dow-cell:nth-child(7) { color: var(--weekend); }
    .cal-dow-cell:not(:nth-child(1)):not(:nth-child(7)) { color: var(--text-sub); }
    .cal-grid { display: grid; grid-template-columns: repeat(7, 1fr); }
    .cal-day { display: flex; flex-direction: column; align-items: center; height: 48px; padding-top: 4px; cursor: default; }
    .day-num { width: 28px; height: 28px; display: flex; align-items: center; justify-content: center; border-radius: 50%; font-size: .8rem; font-weight: 400; flex-shrink: 0; transition: background .12s, color .12s; }
    .cal-day.weekend .day-num { color: var(--weekend); }
    .cal-day:not(.empty):not(.today):hover .day-num { background: var(--blue-pale); color: var(--blue); }
    .cal-day.today .day-num { background: var(--blue); color: #fff !important; font-weight: 700; box-shadow: 0 2px 10px rgba(25,118,210,.38); }
    .cal-day.empty { pointer-events: none; }
    .day-lines { width: 100%; flex: 1; display: flex; flex-direction: column; justify-content: flex-end; gap: 2px; padding-bottom: 3px; overflow: hidden; }
    .week-line { height: 3px; flex-shrink: 0; }

    .week-line-gap {
      height: 3px; flex-shrink: 0; border-radius: 3px;
      background: repeating-linear-gradient(
        90deg, #f59e0b 0, #f59e0b 3px, transparent 3px, transparent 7px
      );
      opacity: .85;
    }

    .gap-badge {
      font-size: .65rem; font-weight: 700; white-space: nowrap;
      background: #fff8e1; border: 1px solid #ffd54f; color: #7c5900;
      border-radius: 10px; padding: 1px 6px;
    }

    .gap-alert {
      display: none; align-items: center; gap: 9px; margin-bottom: 14px;
      padding: 10px 14px; background: #fff8e1;
      border: 1px solid #ffc107; border-radius: 10px;
      font-size: .82rem; color: #7c5900;
    }
    .gap-alert svg { width: 18px; height: 18px; fill: #f59e0b; flex-shrink: 0; }
    .gap-alert strong { font-weight: 700; }
    .gap-alert.visible { display: flex; }

    .ch-gap { font-size: .68rem; color: #f59e0b; margin-left: 4px; }

    .chan-hint {
      display: none; align-items: center; gap: 10px; margin-bottom: 20px;
      padding: 10px 16px; background: var(--blue-pale); border: 1px solid var(--blue-hover);
      border-radius: 10px; font-size: .83rem; font-weight: 500; color: var(--blue-dark);
    }
    .chan-hint.visible { display: flex; }
    .chan-hint strong { font-weight: 700; color: var(--blue); }
    .hint-legend { display: flex; flex-wrap: wrap; gap: 6px; margin-left: auto; }
    .hint-chip { display: flex; align-items: center; gap: 5px; font-size: .72rem; font-weight: 600; background: #fff; border-radius: 20px; padding: 2px 8px 2px 5px; border: 1px solid var(--border); }
    .hint-chip-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }

    /* ── HEATMAP ── */
    .hm-toolbar {
      display: flex; align-items: center; gap: 8px; margin-bottom: 16px; flex-wrap: wrap;
    }
    .hm-metric-btn {
      padding: 7px 16px; font-size: .82rem; font-weight: 600; border-radius: 20px;
      border: 1.5px solid var(--border); background: var(--card-bg); color: var(--text-sub);
      cursor: pointer; transition: all .13s; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
    }
    .hm-metric-btn:hover { border-color: var(--blue-hover); color: var(--blue); background: var(--blue-pale); }
    .hm-metric-btn.active { border-color: var(--blue); color: #fff; background: var(--blue); }

    .hm-info { margin-left: auto; font-size: .75rem; color: var(--text-sub); }

    .hm-empty {
      padding: 64px 24px; text-align: center; color: var(--text-sub);
      font-size: .95rem; background: var(--card-bg);
      border: 1px solid var(--border); border-radius: var(--radius);
    }

    .hm-wrap {
      overflow: auto; border: 1px solid var(--border);
      border-radius: var(--radius); background: var(--card-bg);
      max-height: calc(100vh - 280px);
      box-shadow: var(--shadow);
    }
    .hm-wrap::-webkit-scrollbar { width: 6px; height: 6px; }
    .hm-wrap::-webkit-scrollbar-thumb { background: var(--border); border-radius: 6px; }

    .hm-table { border-collapse: collapse; min-width: 100%; }

    .hm-corner {
      position: sticky; top: 0; left: 0; z-index: 4;
      background: #f0f5fb; min-width: 160px; padding: 10px 14px;
      border-right: 2px solid var(--border); border-bottom: 2px solid var(--border);
      font-size: .68rem; font-weight: 700; text-transform: uppercase;
      letter-spacing: .07em; color: var(--text-sub);
    }

    .hm-head {
      position: sticky; top: 0; z-index: 3;
      background: #f0f5fb; padding: 10px 8px; min-width: 72px;
      border-bottom: 2px solid var(--border); border-right: 1px solid var(--border);
      font-size: .73rem; font-weight: 700; color: var(--blue-dark);
      text-align: center; white-space: nowrap;
    }

    .hm-cat {
      position: sticky; left: 0; z-index: 2;
      background: var(--card-bg); padding: 8px 14px;
      border-right: 2px solid var(--border); border-bottom: 1px solid var(--border);
      font-size: .78rem; font-weight: 600; color: var(--text);
      white-space: nowrap; min-width: 160px;
    }
    .hm-cat-total {
      background: #f5f7fa; color: var(--text-sub); font-weight: 700;
      font-size: .72rem; text-transform: uppercase; letter-spacing: .05em;
    }

    .hm-cell {
      padding: 6px 8px; border-bottom: 1px solid #edf2f8; border-right: 1px solid #edf2f8;
      font-size: .7rem; font-weight: 600; text-align: center; color: var(--text);
      cursor: default; white-space: nowrap;
    }
    .hm-cell-total {
      font-size: .68rem; font-weight: 700; color: #555; border-bottom: 2px solid var(--border);
    }

    .hm-row:nth-child(odd)  { background: #ffffff; }
    .hm-row:nth-child(even) { background: #f4f7fb; }
    .hm-row:hover           { background: #e8f0fb; }
    .hm-row-total           { background: #e4ecf7 !important; }
    .hm-row-total:hover     { background: #d8e5f5 !important; }
    .hm-row:hover .hm-cat  { background: inherit; }
    .hm-row-total .hm-cat  { background: inherit; }
  </style>
</head>
<body>

<!-- SIDEBAR -->
<aside class="sidebar">
  <div class="sidebar-brand">
    <div class="brand-text">
      <h2>Relatorio FTP</h2>
      <span>2025 – 2026</span>
    </div>
  </div>
  <div class="sidebar-body" id="sidebar-body"></div>
  <div class="sidebar-footer">
    <strong id="footer-stat">Canais</strong>
    Selecione um canal para filtrar
  </div>
</aside>

<!-- MAIN -->
<main class="main">
  <div class="page-header">
    <h1>Relatorio FTP</h1>
    <p>Calendario de semanas e heatmap de vendas por categoria</p>
    <div class="today-chip">
      <span class="chip-dot"></span>
      Hoje: <span id="today-label"></span>
    </div>
  </div>

  <div class="tab-bar">
    <button class="tab-btn active" id="tab-cal"  onclick="switchTab('cal')">Calendario</button>
    <button class="tab-btn"        id="tab-heat" onclick="switchTab('heat')">Heatmap de Vendas</button>
  </div>

  <!-- View: Calendario -->
  <div id="view-cal">
    <div class="chan-hint" id="chan-hint">
      <span>Canal: <strong id="hint-chan-name"></strong></span>
      <div class="hint-legend" id="hint-legend"></div>
    </div>
    <div class="gap-alert" id="gap-alert">
      <svg viewBox="0 0 24 24"><path d="M1 21h22L12 2 1 21zm12-3h-2v-2h2v2zm0-4h-2v-4h2v4z"/></svg>
      <span id="gap-alert-text"></span>
    </div>
    <div id="calendar-root"></div>
  </div>

  <!-- View: Heatmap -->
  <div id="view-heat" style="display:none">
    <div class="hm-toolbar">
      <button class="hm-metric-btn active" id="btn-v"   onclick="setMetric('v')">Valor R$</button>
      <button class="hm-metric-btn"        id="btn-vol" onclick="setMetric('vol')">Volume gr/ml</button>
      <button class="hm-metric-btn"        id="btn-u"   onclick="setMetric('u')">Unidades</button>
      <span class="hm-info" id="hm-info"></span>
    </div>
    <div id="heatmap-root"></div>
  </div>
</main>

<script>
// ── Dados embutidos ───────────────────────────────────────────────────────────
const CHANNELS_DATA  = @@CD@@;
const FILE_COLORS    = @@FC@@;
const FILE_LABELS    = @@FL@@;
const CHANNEL_LABELS = @@CL@@;
const SALES_DATA     = @@SD@@;
const FILE_GAPS      = @@FG@@;

// ── Estado ───────────────────────────────────────────────────────────────────
const TODAY = new Date();
let activeTab       = 'cal';
let selectedChannel = null;
let enabledFiles    = new Set();
let activeSalesMetric = 'v';

// ── Constantes ────────────────────────────────────────────────────────────────
const MONTHS_PT  = ['Janeiro','Fevereiro','Marco','Abril','Maio','Junho',
  'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro'];
const DOW_PT     = ['D','S','T','Q','Q','S','S'];
const WEEK_FULL  = ['domingo','segunda-feira','terca-feira','quarta-feira',
  'quinta-feira','sexta-feira','sabado'];

// ── Utils ─────────────────────────────────────────────────────────────────────
function esc(s) {
  return String(s == null ? '' : s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}
function addDays(ds, n) {
  const d = new Date(ds + 'T00:00:00'); d.setDate(d.getDate() + n);
  return d.toISOString().split('T')[0];
}
function datePad(n) { return String(n).padStart(2,'0'); }
function isToday(y,m,d) { return TODAY.getFullYear()===y && TODAY.getMonth()===m && TODAY.getDate()===d; }
function isCurrentMonth(y,m) { return TODAY.getFullYear()===y && TODAY.getMonth()===m; }

function weekLabel(iso) {
  const [,m,d] = iso.split('-');
  return `${d}/${m}`;
}

function fmtShort(v, metric) {
  if (!v || v === 0) return '';
  if (metric === 'v') {
    if (v >= 1e9) return 'R$' + (v/1e9).toFixed(1) + 'B';
    if (v >= 1e6) return 'R$' + (v/1e6).toFixed(1) + 'M';
    if (v >= 1e3) return 'R$' + (v/1e3).toFixed(0) + 'k';
    return 'R$' + v.toFixed(0);
  }
  if (metric === 'vol') {
    if (v >= 1e9) return (v/1e9).toFixed(1) + 'Bg';
    if (v >= 1e6) return (v/1e6).toFixed(1) + 'Mg';
    if (v >= 1e3) return (v/1e3).toFixed(0) + 'kg';
    return v.toFixed(0) + 'g';
  }
  if (v >= 1e9) return (v/1e9).toFixed(1) + 'B';
  if (v >= 1e6) return (v/1e6).toFixed(1) + 'M';
  if (v >= 1e3) return (v/1e3).toFixed(0) + 'k';
  return v.toFixed(0);
}

function fmtFull(v, metric) {
  if (!v || v === 0) return '0';
  const n = v.toLocaleString('pt-BR', {maximumFractionDigits: 2});
  if (metric === 'v')   return 'R$ ' + n;
  if (metric === 'vol') return n + ' gr/ml';
  return n + ' un';
}


// ── Tab switching ─────────────────────────────────────────────────────────────
function switchTab(tab) {
  activeTab = tab;
  document.getElementById('view-cal').style.display  = tab === 'cal'  ? '' : 'none';
  document.getElementById('view-heat').style.display = tab === 'heat' ? '' : 'none';
  document.getElementById('tab-cal').classList.toggle('active',  tab === 'cal');
  document.getElementById('tab-heat').classList.toggle('active', tab === 'heat');
  if (tab === 'heat') renderHeatmap();
}

// ── Heatmap ───────────────────────────────────────────────────────────────────
function setMetric(m) {
  activeSalesMetric = m;
  ['v','vol','u'].forEach(k => {
    document.getElementById('btn-'+k).classList.toggle('active', k === m);
  });
  renderHeatmap();
}

function renderHeatmap() {
  const root = document.getElementById('heatmap-root');
  const info = document.getElementById('hm-info');

  if (!selectedChannel) {
    root.innerHTML = '<div class="hm-empty">Selecione um canal na sidebar para ver o heatmap de vendas.</div>';
    info.textContent = '';
    return;
  }

  const chSales = SALES_DATA[selectedChannel] || {};
  const cats = Object.keys(chSales).sort((a, b) => {
    const aT = a.toLowerCase() === 'total', bT = b.toLowerCase() === 'total';
    if (aT && !bT) return 1;
    if (!aT && bT) return -1;
    return a.localeCompare(b, 'pt-BR');
  });

  const weeksSet = new Set();
  for (const weekData of Object.values(chSales)) {
    for (const w of Object.keys(weekData)) weeksSet.add(w);
  }
  const weeks = [...weeksSet].sort();

  if (!cats.length || !weeks.length) {
    root.innerHTML = '<div class="hm-empty">Sem dados de vendas para este canal.</div>';
    info.textContent = '';
    return;
  }

  const metricNames = { v: 'Valor R$', vol: 'Volume gr/ml', u: 'Unidades' };
  info.textContent = `${CHANNEL_LABELS[selectedChannel] || selectedChannel} | ${cats.length} categorias | ${weeks.length} semanas | ${metricNames[activeSalesMetric]}`;

  let html = '<div class="hm-wrap"><table class="hm-table"><thead><tr>';
  html += `<th class="hm-corner">Categoria</th>`;
  for (const w of weeks) {
    html += `<th class="hm-head">${esc(weekLabel(w))}<br><span style="font-weight:400;font-size:.65rem;opacity:.7">${w.slice(0,4)}</span></th>`;
  }
  html += '</tr></thead><tbody>';

  for (const cat of cats) {
    const isTotal = cat.toLowerCase() === 'total';
    html += `<tr class="${isTotal ? 'hm-row-total' : 'hm-row'}">`;
    html += `<td class="hm-cat ${isTotal ? 'hm-cat-total' : ''}">${esc(cat)}</td>`;
    for (const w of weeks) {
      const vals = (chSales[cat] || {})[w] || {};
      const v = vals[activeSalesMetric] || 0;
      const lbl = fmtShort(v, activeSalesMetric);
      const tip = `${cat} | ${w}\n${metricNames[activeSalesMetric]}: ${fmtFull(v, activeSalesMetric)}`;
      html += `<td class="hm-cell ${isTotal ? 'hm-cell-total' : ''}" title="${esc(tip)}">${esc(lbl)}</td>`;
    }
    html += '</tr>';
  }

  html += '</tbody></table></div>';
  root.innerHTML = html;
}

// ── Calendario ────────────────────────────────────────────────────────────────
function buildMonthCard(year, month) {
  const card = document.createElement('div');
  card.className = 'cal-card' + (isCurrentMonth(year, month) ? ' is-current' : '');

  const top = document.createElement('div');
  top.className = 'cal-card-top';
  const lbl = document.createElement('span');
  lbl.className = 'cal-month-label';
  lbl.textContent = MONTHS_PT[month];
  top.appendChild(lbl);
  if (isCurrentMonth(year, month)) {
    const b = document.createElement('span');
    b.className = 'cal-badge'; b.textContent = 'Atual';
    top.appendChild(b);
  }
  card.appendChild(top);

  const dowRow = document.createElement('div');
  dowRow.className = 'cal-dow-row';
  DOW_PT.forEach(l => {
    const c = document.createElement('div');
    c.className = 'cal-dow-cell'; c.textContent = l;
    dowRow.appendChild(c);
  });
  card.appendChild(dowRow);

  const grid = document.createElement('div');
  grid.className = 'cal-grid';
  const startDow = new Date(year, month, 1).getDay();
  const numDays  = new Date(year, month + 1, 0).getDate();

  for (let i = 0; i < startDow; i++) {
    const e = document.createElement('div'); e.className = 'cal-day empty'; grid.appendChild(e);
  }
  for (let d = 1; d <= numDays; d++) {
    const dow  = (startDow + d - 1) % 7;
    const isTd = isToday(year, month, d);
    const ds   = `${year}-${datePad(month + 1)}-${datePad(d)}`;
    const cell = document.createElement('div');
    cell.className = 'cal-day' + ((dow === 0 || dow === 6) ? ' weekend' : '') + (isTd ? ' today' : '');
    cell.dataset.date = ds;
    if (isTd) cell.setAttribute('data-today', '');
    const num = document.createElement('span');
    num.className = 'day-num'; num.textContent = d;
    cell.appendChild(num);
    const lines = document.createElement('div');
    lines.className = 'day-lines';
    cell.appendChild(lines);
    grid.appendChild(cell);
  }
  card.appendChild(grid);
  return card;
}

function buildYearSection(year) {
  const sec = document.createElement('div');
  sec.className = 'year-section';
  sec.innerHTML = `<div class="year-heading"><span class="year-num">${year}</span><div class="year-line"></div></div>`;
  const g = document.createElement('div');
  g.className = 'months-grid';
  for (let m = 0; m < 12; m++) g.appendChild(buildMonthCard(year, m));
  sec.appendChild(g);
  return sec;
}

function applyCalendar() {
  document.querySelectorAll('.day-lines').forEach(el => { el.innerHTML = ''; });
  if (!selectedChannel) return;
  const chData = CHANNELS_DATA[selectedChannel] || {};

  // ── Linhas normais (semanas com dados) ──
  const coverage = {};
  for (const [fileKey, dates] of Object.entries(chData)) {
    if (!enabledFiles.has(fileKey)) continue;
    const color = FILE_COLORS[fileKey] || '#999';
    for (const monday of dates) {
      for (let i = 0; i < 7; i++) {
        const ds = addDays(monday, i);
        if (!coverage[ds]) coverage[ds] = [];
        if (!coverage[ds].some(e => e.file === fileKey))
          coverage[ds].push({ file: fileKey, color, weekPos: i });
      }
    }
  }
  for (const [dateStr, entries] of Object.entries(coverage)) {
    const cell = document.querySelector(`[data-date="${dateStr}"]`);
    if (!cell) continue;
    const linesEl = cell.querySelector('.day-lines');
    if (!linesEl) continue;
    for (const { file, color, weekPos } of entries) {
      const line = document.createElement('div');
      line.className = 'week-line';
      line.style.background = color;
      line.title = FILE_LABELS[file] || file;
      if (weekPos === 0)      { line.style.borderRadius = '3px 0 0 3px'; line.style.marginLeft = '3px'; }
      else if (weekPos === 6) { line.style.borderRadius = '3px'; line.style.marginLeft = '2px'; line.style.marginRight = '3px'; }
      else                    { line.style.borderRadius = '0'; }
      linesEl.appendChild(line);
    }
  }

  // ── Linhas de gap (semanas faltantes) — marcadas so na segunda-feira ──
  for (const [fileKey, gaps] of Object.entries(FILE_GAPS)) {
    if (!enabledFiles.has(fileKey)) continue;
    if (!chData[fileKey]) continue;
    const label = FILE_LABELS[fileKey] || fileKey;
    for (const monday of gaps) {
      const cell = document.querySelector(`[data-date="${monday}"]`);
      if (!cell) continue;
      const linesEl = cell.querySelector('.day-lines');
      if (!linesEl) continue;
      // Evita duplicar se outro arquivo ja marcou este gap
      if (linesEl.querySelector('.week-line-gap[data-file="' + fileKey + '"]')) continue;
      const line = document.createElement('div');
      line.className = 'week-line-gap';
      line.dataset.file = fileKey;
      line.title = `⚠ Semana faltante: ${label}`;
      line.style.marginLeft = '3px';
      linesEl.appendChild(line);
    }
  }
}

function updateGapAlert() {
  const alertEl = document.getElementById('gap-alert');
  const textEl  = document.getElementById('gap-alert-text');
  if (!selectedChannel) { alertEl.classList.remove('visible'); return; }

  const chData = CHANNELS_DATA[selectedChannel] || {};
  let totalGaps = 0;
  const filesWithGaps = [];
  for (const fileKey of Object.keys(chData)) {
    if (!enabledFiles.has(fileKey)) continue;
    const gaps = FILE_GAPS[fileKey] || [];
    if (gaps.length > 0) {
      totalGaps += gaps.length;
      filesWithGaps.push({ label: FILE_LABELS[fileKey] || fileKey, count: gaps.length });
    }
  }

  if (totalGaps === 0) { alertEl.classList.remove('visible'); return; }

  const detail = filesWithGaps.map(f => `${esc(f.label)} (${f.count})`).join(', ');
  textEl.innerHTML = `<strong>${totalGaps} semana(s) faltando</strong> no intervalo: ${detail}`;
  alertEl.classList.add('visible');
}

function updateHintBar() {
  const bar  = document.getElementById('chan-hint');
  const name = document.getElementById('hint-chan-name');
  const leg  = document.getElementById('hint-legend');
  if (!selectedChannel) { bar.classList.remove('visible'); updateGapAlert(); return; }
  bar.classList.add('visible');
  name.textContent = CHANNEL_LABELS[selectedChannel] || selectedChannel;
  leg.innerHTML = '';
  const chData = CHANNELS_DATA[selectedChannel] || {};
  for (const fileKey of Object.keys(chData)) {
    if (!enabledFiles.has(fileKey)) continue;
    const chip = document.createElement('span');
    chip.className = 'hint-chip';
    chip.innerHTML = `<span class="hint-chip-dot" style="background:${FILE_COLORS[fileKey]||'#999'}"></span>${esc(FILE_LABELS[fileKey]||fileKey)}`;
    leg.appendChild(chip);
  }
  updateGapAlert();
}

// ── Sidebar ───────────────────────────────────────────────────────────────────
function showChannelList() {
  selectedChannel = null;
  enabledFiles.clear();
  applyCalendar();
  updateHintBar();
  if (activeTab === 'heat') renderHeatmap();

  const body = document.getElementById('sidebar-body');
  body.innerHTML = '';
  const sec = document.createElement('div');
  sec.className = 'sb-section'; sec.textContent = 'Canais';
  body.appendChild(sec);

  for (const [ch, files] of Object.entries(CHANNELS_DATA)) {
    const fileKeys   = Object.keys(files);
    const hasGaps    = fileKeys.some(k => (FILE_GAPS[k] || []).length > 0);
    const gapIcon    = hasGaps ? `<span class="ch-gap" title="Alguns arquivos possuem semanas faltando">⚠</span>` : '';
    const item = document.createElement('div');
    item.className = 'sb-channel';
    item.innerHTML = `<span class="ch-dot"></span><span>${esc(CHANNEL_LABELS[ch]||ch)}</span>${gapIcon}<span class="ch-count">${fileKeys.length}</span>`;
    item.addEventListener('click', () => showChannelFiles(ch));
    body.appendChild(item);
  }

  document.getElementById('footer-stat').textContent =
    `${Object.keys(CHANNELS_DATA).length} canal(is)`;
}

function showChannelFiles(channelName) {
  selectedChannel = channelName;
  const chData = CHANNELS_DATA[channelName] || {};
  enabledFiles = new Set(Object.keys(chData));
  applyCalendar();
  updateHintBar();
  if (activeTab === 'heat') renderHeatmap();

  const body = document.getElementById('sidebar-body');
  body.innerHTML = '';

  const back = document.createElement('div');
  back.className = 'sb-back';
  back.innerHTML = `<svg viewBox="0 0 24 24"><path d="M20 11H7.83l5.59-5.59L12 4l-8 8 8 8 1.41-1.41L7.83 13H20v-2z"/></svg> Canais`;
  back.addEventListener('click', showChannelList);
  body.appendChild(back);

  const hdr = document.createElement('div');
  hdr.className = 'sb-chan-header';
  hdr.textContent = CHANNEL_LABELS[channelName] || channelName;
  body.appendChild(hdr);

  const ctrl = document.createElement('div');
  ctrl.className = 'sb-controls';
  ['Todos', 'Nenhum'].forEach((label, idx) => {
    const btn = document.createElement('span');
    btn.className = 'sb-btn'; btn.textContent = label;
    btn.addEventListener('click', () => {
      enabledFiles = idx === 0 ? new Set(Object.keys(chData)) : new Set();
      renderFileList();
      applyCalendar();
      updateHintBar();
    });
    ctrl.appendChild(btn);
  });
  body.appendChild(ctrl);

  const div = document.createElement('div'); div.className = 'sb-div'; body.appendChild(div);
  const listEl = document.createElement('div'); listEl.id = 'sb-file-list'; body.appendChild(listEl);

  function renderFileList() {
    listEl.innerHTML = '';
    for (const fileKey of Object.keys(chData)) {
      const color   = FILE_COLORS[fileKey] || '#999';
      const label   = FILE_LABELS[fileKey] || fileKey;
      const isOn    = enabledFiles.has(fileKey);
      const gapCnt  = (FILE_GAPS[fileKey] || []).length;
      const gapHtml = gapCnt > 0
        ? `<span class="gap-badge" title="${gapCnt} semana(s) faltando no intervalo">⚠ ${gapCnt}</span>`
        : '';
      const item = document.createElement('div');
      item.className = 'sb-file' + (isOn ? '' : ' off');
      item.innerHTML = `
        <span class="file-swatch" style="background:${color}"></span>
        <span class="file-label" title="${esc(fileKey)}">${esc(label)}</span>
        ${gapHtml}
        <span class="file-check ${isOn ? 'on' : ''}"></span>`;
      item.addEventListener('click', () => {
        if (enabledFiles.has(fileKey)) enabledFiles.delete(fileKey);
        else enabledFiles.add(fileKey);
        renderFileList();
        applyCalendar();
        updateHintBar();
      });
      listEl.appendChild(item);
    }
  }
  renderFileList();

  document.getElementById('footer-stat').textContent =
    `${Object.keys(chData).length} arquivo(s)`;
}

// ── Init ──────────────────────────────────────────────────────────────────────
const calRoot = document.getElementById('calendar-root');
calRoot.appendChild(buildYearSection(2025));
calRoot.appendChild(buildYearSection(2026));

document.getElementById('today-label').textContent =
  `${WEEK_FULL[TODAY.getDay()]}, ${TODAY.getDate()} de ${MONTHS_PT[TODAY.getMonth()]} de ${TODAY.getFullYear()}`;

showChannelList();

setTimeout(() => {
  document.querySelector('[data-today]')?.closest('.cal-card')
    ?.scrollIntoView({ behavior: 'smooth', block: 'center' });
}, 300);
</script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print('Lendo arquivos de resultado_v2...')

    if not os.path.isdir(BASE_DIR):
        print(f'ERRO: Pasta nao encontrada: {BASE_DIR}')
        return

    channels_data: Dict[str, Dict] = {}
    all_filtrado_rows: List[Dict]   = []

    for canal in sorted(os.listdir(BASE_DIR)):
        canal_dir = os.path.join(BASE_DIR, canal)
        if not os.path.isdir(canal_dir):
            continue

        unicos   = read_unicos(canal_dir)
        filtrado = read_filtrado_canal(canal, canal_dir)

        if unicos:
            channels_data[canal] = unicos
        all_filtrado_rows.extend(filtrado)
        print(f'  {canal}: {len(unicos)} arq(unicos) | {len(filtrado)} linhas(filtrado)')

    n_ch    = len(channels_data)
    n_files = sum(len(f) for f in channels_data.values())
    n_weeks = sum(len(d) for ch in channels_data.values() for d in ch.values())
    print(f'Total: {n_ch} canais | {n_files} arquivos | {n_weeks} semanas | {len(all_filtrado_rows)} linhas filtrado')

    # Estruturas para o HTML
    all_files   = sorted({f for ch in channels_data.values() for f in ch})
    file_colors = {f: PALETTE[i % len(PALETTE)] for i, f in enumerate(all_files)}
    file_labels = {f: fmt_file_date(f) for f in all_files}
    chan_labels  = {ch: fmt_channel(ch) for ch in channels_data}
    sales_data   = build_sales_data(all_filtrado_rows)

    n_cats  = sum(len(cats) for cats in sales_data.values())
    n_cells = sum(len(w) for cats in sales_data.values() for w in cats.values())
    print(f'  Heatmap: {n_cats} categorias | {n_cells} celulas de dados')

    # Serializar JSON
    CD = json.dumps(channels_data, ensure_ascii=False, separators=(',', ':'))
    FC = json.dumps(file_colors,   ensure_ascii=False, separators=(',', ':'))
    FL = json.dumps(file_labels,   ensure_ascii=False, separators=(',', ':'))
    CL = json.dumps(chan_labels,   ensure_ascii=False, separators=(',', ':'))
    SD = json.dumps(sales_data,    ensure_ascii=False, separators=(',', ':'))
    FG = '{}'

    html = (HTML_TEMPLATE
            .replace('@@CD@@', CD)
            .replace('@@FC@@', FC)
            .replace('@@FL@@', FL)
            .replace('@@CL@@', CL)
            .replace('@@SD@@', SD)
            .replace('@@FG@@', FG))

    os.makedirs(os.path.dirname(OUTPUT_HTML), exist_ok=True)
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'HTML gerado em: {OUTPUT_HTML}')

    # Gerar Excel
    if all_filtrado_rows:
        write_excel(all_filtrado_rows)
    else:
        print('  Nenhum dado filtrado encontrado. Excel nao gerado.')


if __name__ == '__main__':
    main()
